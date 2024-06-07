# -*- coding: utf-8 -*-

import openpyxl
from math import ceil
from openpyxl.styles import Font


def weightsCalcMaxPerson(pathToFile, weightTick, maxKartWeight):
    participantsFile = openpyxl.load_workbook(str(pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))] + "/participants_weights.xlsx"))
    participantsSheet = participantsFile["Sheet"]
    pos = 2
    maxParticipantWeight = float(0)
    while participantsSheet.cell(row = pos, column = 2).value != None:
        if participantsSheet.cell(row = pos, column = 2).value > maxParticipantWeight:
            maxParticipantWeight = participantsSheet.cell(row = pos, column = 2).value
        pos += 1
    pos = 2
    participantsSheet.cell(row = 1, column = 3).font = Font(size = 14, bold = True)
    participantsSheet.cell(row = 1, column = 3).value = str("Довес")
    while participantsSheet.cell(row = pos, column = 2).value != None:
        sheetWeightWrite = participantsSheet.cell(row = pos, column = 3)
        weightCount = participantsSheet.cell(row = pos, column = 2).value
        while weightCount + weightTick < maxParticipantWeight and weightCount - participantsSheet.cell(row = pos, column = 2).value + weightTick <= maxKartWeight:
            weightCount += weightTick
        sheetWeightWrite.value = weightCount - participantsSheet.cell(row = pos, column = 2).value
        pos += 1
        
    participantsFile.save(str(pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))] + "/participants_weights.xlsx"))
    
def weightsCalcFixed(pathToFile, weightTick, maxKartWeight, fixedWeight):
    participantsFile = openpyxl.load_workbook("%s/participants_weights.xlsx" % (pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))]))
    participantsSheet = participantsFile["Sheet"]    
    pos = 2
    participantsSheet.cell(row = 1, column = 3).font = Font(size = 14, bold = True)
    participantsSheet.cell(row = 1, column = 3).value = str("Довес")
    while participantsSheet.cell(row = pos, column = 2).value != None:
        sheetWeightWrite = participantsSheet.cell(row = pos, column = 3)
        weightCount = participantsSheet.cell(row = pos, column = 2).value
        while weightCount + weightTick < fixedWeight and weightCount - participantsSheet.cell(row = pos, column = 2).value + weightTick <= maxKartWeight:
            weightCount += weightTick
        sheetWeightWrite.value = weightCount - participantsSheet.cell(row = pos, column = 2).value
        pos += 1
    participantsFile.save(str(pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))] + "/participants_weights.xlsx"))