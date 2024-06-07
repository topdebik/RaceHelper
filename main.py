# -*- coding: utf-8 -*-

import sys
import os
from PyQt5 import QtWidgets, QtGui
from openpyxl import load_workbook
from errors import sameParticipantsError
import mainWindow
import toolGeneratePlacesWindow
import toolGenerateWeightsWindow
import toolGenerateResultsWindow
import places_func
import results_func
import weights_func


class mainWindowApp(QtWidgets.QMainWindow, mainWindow.Ui_MainWindow):   
    def __init__(self):
        try:
            super().__init__()         
            self.setupUi(self)
            self.setWindowTitle("Автоспортивный помощник")
            self.menuHelpActionOpenHelp.triggered.connect(self.showHelp)
            self.menuProgramActionQuit.triggered.connect(QtWidgets.qApp.quit)
            self.btnChooseCurrentPath.clicked.connect(self.browseFolderAndShowParticipants)
            self.btnGeneratePlaces.clicked.connect(self.showToolGeneratePlacesWindow)
            self.btnGenerateWeights.clicked.connect(self.showToolGenerateWeightsWindow)
            self.btnGenerateResults.clicked.connect(self.showToolGenerateResultsWindow)
        except Exception as e:
            showErrorMessage(e)
    def showHelp(self):
        try:
            os.startfile("Help.pdf")
        except Exception as e:
            showErrorMessage(e)        
    def browseFolderAndShowParticipants(self):
        try:
            global directory
            directory = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel Files (*.xlsx);;All Files (*)")
            if directory[0] != "":
                self.textChosenCurrentPath.setText(str(directory[0]))
                self.listParticipants.clear()
                for participant in countParticipants(str(directory[0])):
                    self.listParticipants.addItem(participant)
            else:
                self.textChosenCurrentPath.setText("Выберите файл с участниками")
                self.listParticipants.clear()
        except Exception as e:
            showErrorMessage(e)
    def onSelected(self, text):
        try:
            global generateMethod
            generateMethod = text
        except Exception as e:
            showErrorMessage(e)        
    def valueChangePlaces(self):
        try:
            global maxKarts, minKarts, runsPerParticipant
            maxKarts = self.window.boxMaxKarts.value()
            minKarts = self.window.boxMinKarts.value()
            runsPerParticipant = self.window.boxRunsPerParticipant.value()
        except Exception as e:
            showErrorMessage(e)
    def valueChangeWeights(self):
        try:
            global weightTick, maxKartWeight, weightGoal
            weightTick = self.window.boxWeightTick.value()
            maxKartWeight = self.window.boxMaxKartWeight.value()
            weightGoal = self.window.boxWeightGoal.value()
        except Exception as e:
            showErrorMessage(e)
    def showCompletedAlert(self):
        try:
            self.message = QtWidgets.QMessageBox
            if "directory" not in globals() or directory[0] == "":
                self.message.about(self, "Генератор","Файл с участниками не выбран")
            else:
                if "Results" in str(self.window) and "resultsDirectory" not in globals():
                    self.message.about(self, "Генератор","Файл с конфигурацией очков не выбран")
                else:
                    if "resultsNotFound" in globals():
                        global resultsNotFound
                        del resultsNotFound
                        self.message.about(self, "Генератор","Результаты по заездам не были найдены")
                    else:
                        self.message.about(self, "Генератор","Готово!")
            self.window.close()
        except Exception as e:
            showErrorMessage(e)       
    def browseFolderAndSelectResultsConfig(self):
        try:
            global resultsDirectory
            resultsDirectory = QtWidgets.QFileDialog.getOpenFileName(self.window, "Выберите файл", "", "Text Files (*.txt);;All Files (*)")
            if resultsDirectory[0] != "":
                self.window.textChosenCurrentPath.setText(str(resultsDirectory[0]))
        except Exception as e:
            showErrorMessage(e)    
    def showToolGeneratePlacesWindow(self):
        def places_func_activate():
            try:
                if "directory" in globals() and directory[0] != "":
                    if "generateMethod" not in globals():
                        global generateMethod
                        generateMethod = "Случайно"
                    if generateMethod == "Случайно":
                        places_func.placeCalcRandom(str(directory[0]), maxKarts, runsPerParticipant, minKarts = minKarts)
                    elif generateMethod == "По алфавиту":
                        places_func.placeCalcAlphabet(str(directory[0]), maxKarts, runsPerParticipant, minKarts = minKarts)
                    elif generateMethod == "По алфавиту, в обратном порядке":
                        places_func.placeCalcAlphabet(str(directory[0]), maxKarts, runsPerParticipant, minKarts = minKarts, revCalc = True)
            except Exception as e:
                showErrorMessage(e)           
        try:
            self.window = toolGeneratePlacesWindowApp()
            self.window.show()
            self.window.btnGenerate.clicked.connect(places_func_activate)
            self.window.btnGenerate.clicked.connect(self.showCompletedAlert)
            self.window.boxGenerateMethod.activated[str].connect(self.onSelected)
            self.window.boxMaxKarts.valueChanged.connect(self.valueChangePlaces)
            self.window.boxMinKarts.valueChanged.connect(self.valueChangePlaces)
            self.window.boxRunsPerParticipant.valueChanged.connect(self.valueChangePlaces)
        except Exception as e:
            showErrorMessage(e)      
    def showToolGenerateWeightsWindow(self):
        def weights_func_activate():
            try:
                if "directory" in globals() and directory[0] != "":
                    if "generateMethod" not in globals():
                        global generateMethod
                        generateMethod = "До максимального веса"
                    if generateMethod == "До максимального веса":
                        weights_func.weightsCalcMaxPerson(str(directory[0]), weightTick, maxKartWeight)
                    elif generateMethod == "До указанного веса":
                        weights_func.weightsCalcFixed(str(directory[0]), weightTick, maxKartWeight, weightGoal)
            except Exception as e:
                showErrorMessage(e)          
        try:
            self.window = toolGenerateWeightsWindowApp()
            self.window.show()
            self.window.btnGenerate.clicked.connect(weights_func_activate)
            self.window.btnGenerate.clicked.connect(self.showCompletedAlert)
            self.window.boxGenerateMethod.activated[str].connect(self.onSelected)
            self.window.boxWeightTick.valueChanged.connect(self.valueChangeWeights)
            self.window.boxMaxKartWeight.valueChanged.connect(self.valueChangeWeights)
            self.window.boxWeightGoal.valueChanged.connect(self.valueChangeWeights)
        except Exception as e:
            showErrorMessage(e)
    def showToolGenerateResultsWindow(self):
        def results_func_activate():
            try:
                if "directory" in globals() and directory[0] != "" and "resultsDirectory" in globals():
                    results_func.resultsCalcCustom(str(directory[0]), str(resultsDirectory[0]))
            except Exception as e:
                if str(e) == "No results were found":
                    global resultsNotFound
                    resultsNotFound = True
                else:
                    showErrorMessage(e)
        try:
            self.window = toolGenerateResultsWindowApp()
            self.window.show()
            self.window.btnChooseCurrentPath.clicked.connect(self.browseFolderAndSelectResultsConfig)
            self.window.btnGenerate.clicked.connect(results_func_activate)
            self.window.btnGenerate.clicked.connect(self.showCompletedAlert)
        except Exception as e:
            showErrorMessage(e)   
        
class toolGeneratePlacesWindowApp(QtWidgets.QMainWindow, toolGeneratePlacesWindow.Ui_toolGeneratePlacesWindow):
    def __init__(self):
        try:
            super().__init__()
            self.setupUi(self)
            self.setWindowTitle("Настройки генератора стартовых мест")
            self.setWindowModality(2)
        except Exception as e:
            showErrorMessage(e)        
class toolGenerateWeightsWindowApp(QtWidgets.QMainWindow, toolGenerateWeightsWindow.Ui_toolGenerateWeightsWindow):
    def __init__(self):
        try:
            super().__init__()
            self.setupUi(self)
            self.setWindowTitle("Настройки генератора довесов")
            self.setWindowModality(2)
        except Exception as e:
            showErrorMessage(e)  
class toolGenerateResultsWindowApp(QtWidgets.QMainWindow, toolGenerateResultsWindow.Ui_toolGenerateResultsWindow):
    def __init__(self):
        try:
            super().__init__()
            self.setupUi(self)
            self.setWindowTitle("Настройки генератора результатов")
            self.setWindowModality(2)
        except Exception as e:
            showErrorMessage(e)     

def main():
    try:
        app = QtWidgets.QApplication(sys.argv)
        window = mainWindowApp()
        window.show()
        app.exec_()
    except Exception as e:
        showErrorMessage(e)    
               
def countParticipants(pathToFile):
    try:
        participantsFile = load_workbook(pathToFile)
        participantsSheet = participantsFile["Sheet"]
        pos = 2
        participantsTab = list()
        while participantsSheet.cell(row = pos, column = 1).value != None:
            if participantsSheet.cell(row = pos, column = 1).value in participantsTab:
                raise sameParticipantsError           
            participantsTab.append(str(participantsSheet.cell(row = pos, column = 1).value))     
            pos += 1
        return participantsTab
    except Exception as e:
        showErrorMessage(e)

def showErrorMessage(e):
    errDialog = QtWidgets.QMessageBox()
    errDialog.setIcon(QtWidgets.QMessageBox.Critical)
    errDialog.setText("Error")
    errDialog.setInformativeText("Error occured: \042%s\042. Application will now close." % str(e))
    errDialog.setWindowTitle("Error")
    errDialog.exec_()
    exit()     

if __name__ == "__main__":
    main()