# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from errors import noResultsError


def lrindex(alist, value):
    return len(alist) - alist[-1 : : -1].index(value) - 1

def parseResults(resultsTab, points, pathToFile):
    participantsFile = Workbook()
    participantsSheet = participantsFile.active
    pos = 1
    for i in range(len(resultsTab)):
        participantsSheet.cell(row = pos, column = 1).value = (i + 1)
        participantsSheet.cell(row = pos, column = 2).value = resultsTab[i]
        if len(points) - 1 < i:
            participantsSheet.cell(row = pos, column = 3).value = 0
        else:
            participantsSheet.cell(row = pos, column = 3).value = points[i]
        pos += 1
    participantsFile.save(str(pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))] + "/participants_results.xlsx"))

def sortByPlaces(sameSumsParticipantsList):
    output = list()
    places = list()
    for i in range(len(sameSumsParticipantsList)):
        for f in range(len(sameSumsParticipantsList[i][1])):
            places.append(sameSumsParticipantsList[i][1][f])
    places = sorted(list(set(places)))
    participantsPlaces = list()
    for i in range(len(sameSumsParticipantsList)):
        participantsPlaces.append([sameSumsParticipantsList[i][0], [sameSumsParticipantsList[i][1].count(place) for place in places]])
    participantsPlaces.sort(key = lambda x: tuple(x), reverse = True)
    for i in range(len(participantsPlaces)):
        output.append(sameSumsParticipantsList[[sameSumsParticipantsList[x][0] for x in range(len(sameSumsParticipantsList))].index(participantsPlaces[i][0])])
    return output    

def resultsCalcCustom(pathToFile, pathToCustomConfig):
    participantsFile = load_workbook(str(pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))] + "/participants_places.xlsx"))
    participantsSheet = participantsFile["Sheet"]
    customPoints = list()
    with open(str(pathToCustomConfig), "r") as File:
        for line in File:
            customPoints.append(int(line))
    pos = 1
    participants = list()
    participantsPlaces = list()
    while participantsSheet.cell(row = pos, column = 2).value != None or participantsSheet.cell(row = pos + 1, column = 2).value != None:
        if participantsSheet.cell(row = pos, column = 2).value != None and "Заезд" not in participantsSheet.cell(row = pos, column = 2).value and participantsSheet.cell(row = pos, column = 2).value not in participants and "Mr. Nobody" not in participantsSheet.cell(row = pos, column = 2).value:
            participants.append(participantsSheet.cell(row = pos, column = 2).value)
        pos += 1
    for i in range(len(participants)):
        pos = 1
        currentParticipantPlaces = list()
        while participantsSheet.cell(row = pos, column = 2).value != None or participantsSheet.cell(row = pos + 1, column = 2).value != None:
            if participantsSheet.cell(row = pos, column = 2).value != None and participantsSheet.cell(row = pos, column = 2).value == participants[i]:
                if participantsSheet.cell(row = pos, column = 3).value == None:
                    raise noResultsError
                currentParticipantPlaces.append(participantsSheet.cell(row = pos, column = 3).value)
            pos += 1
        participantsPlaces.append([participants[i], currentParticipantPlaces])
    for i in range(len(participantsPlaces)):
        if "Mr. Nobody" in participantsPlaces[i][0]:
            participantsPlaces.remove(participantsPlaces[i])
    participantsPlaces.sort(key = lambda x: sum(x[1]))
    participantsPlacesSums = list()
    for i in range(len(participantsPlaces)):
        participantsPlacesSums.append(sum(participantsPlaces[i][1]))
    participantsPlacesUniqueSums = list(set(participantsPlacesSums))
    for sums in participantsPlacesUniqueSums:
        if participantsPlacesSums.count(sums) > 1:
            sameSumsParticipantsList = list()
            for i in range(participantsPlacesSums.index(sums), lrindex(participantsPlacesSums, sums) + 1):
                sameSumsParticipantsList.append(participantsPlaces[i])
            sameSumsParticipantsList = sortByPlaces(sameSumsParticipantsList)
            for i in range(participantsPlacesSums.index(sums), lrindex(participantsPlacesSums, sums) + 1):
                participantsPlaces[i] = sameSumsParticipantsList[i - participantsPlacesSums.index(sums)]
    placesTab = list()
    for i in range(len(participantsPlaces)):
        placesTab.append(participantsPlaces[i][0])
    parseResults(placesTab, customPoints, pathToFile)