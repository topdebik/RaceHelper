# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import random
from math import ceil
from errors import runsAmountError, sameParticipantsError, tooManyNobodysError


def parsePlaces(placesTab, pathToFile):
    participantsFile = Workbook()
    participantsSheet = participantsFile.active
    pos = 1
    for i in range(len(placesTab)):
        if pos != 1:
            pos += 1
        participantsSheet.cell(row = pos, column = 2).font = Font(size = 14, bold = True)
        participantsSheet.cell(row = pos, column = 2).value = str("Заезд %s") % (i + 1)
        pos += 1
        for f in range(len(placesTab[i])):
            participantsSheet.cell(row = pos, column = 1).value = (f + 1)
            participantsSheet.cell(row = pos, column = 2).value = str(placesTab[i][f])
            pos += 1
    participantsFile.save(str(pathToFile[:pathToFile.rfind("/", 0, len(pathToFile))] + "/participants_places.xlsx"))

def placeCalcRandom(pathToFile, availableKarts, runsPerParticipant, minKarts = 0):
    if minKarts == 0:
        minKarts = availableKarts
    participantsFile = load_workbook(pathToFile)
    participantsSheet = participantsFile["Sheet"]
    pos = 2
    participantsNumber = 0
    if runsPerParticipant % 2 != 0 and runsPerParticipant != 1:
        raise runsAmountError
    participantsTab = list()
    while participantsSheet.cell(row = pos, column = 1).value != None:
        if participantsSheet.cell(row = pos, column = 1).value in participantsTab:
            raise sameParticipantsError           
        participantsNumber += 1
        participantsTab.append(str(participantsSheet.cell(row = pos, column = 1).value))     
        pos += 1
    karts = 0
    nobodys = 0
    while karts < minKarts:
        karts = availableKarts
        while ((participantsNumber + nobodys) % karts != 0):
            karts -= 1
        if karts < minKarts:
            nobodys += 1
        if nobodys > 5:
            raise tooManyNobodysError
    if nobodys > 0:
        for i in range(nobodys):
            participantsNumber += 1
            participantsTab.append(str("Mr. Nobody %s") % str(i + 1))      
    groups = participantsNumber // karts
    placesTab = list()
    for i in range(runsPerParticipant):
        if i % 2 == 0:
            currentRunTab = list(participantsTab)
            for f in range(groups):
                groupTab = list()
                for j in range(karts):
                    currentParticipant = random.choice(currentRunTab)
                    groupTab.append(currentParticipant)
                    currentRunTab.remove(currentParticipant)
                placesTab.append(groupTab)
        elif i % 2 != 0:
            currentRunTab = [[karts * groups for groups in range(0, groups)] for karts in range(0, karts)]
            for f in range(groups):
                for j in range(karts):
                    currentRunTab[j][f] = str(placesTab[((i * groups) - groups) + f][j])
            for f in range(groups):
                groupTab = list()
                for j in range(-1, (karts * -1) - 1, -1):
                    currentParticipant = random.choice(currentRunTab[j])
                    groupTab.append(currentParticipant)
                    currentRunTab[j].remove(currentParticipant)
                placesTab.append(groupTab)
    parsePlaces(placesTab, pathToFile)

def placeCalcAlphabet(pathToFile, availableKarts, runsPerParticipant, minKarts = 0, revCalc = False):
    if minKarts == 0:
        minKarts = availableKarts    
    participantsFile = load_workbook(pathToFile)
    participantsSheet = participantsFile["Sheet1"]
    pos = 2
    participantsNumber = 0
    if runsPerParticipant % 2 != 0 and runsPerParticipant != 1:
        raise runsAmountError
    participantsTab = list()
    while participantsSheet.cell(row = pos, column = 1).value != None:
        if participantsSheet.cell(row = pos, column = 1).value in participantsTab:
            raise sameParticipantsError           
        participantsNumber += 1
        participantsTab.append(str(participantsSheet.cell(row = pos, column = 1).value))     
        pos += 1
    participantsTab.sort(reverse = revCalc)
    karts = 0
    nobodys = 0
    while karts < minKarts:
        karts = availableKarts
        while ((participantsNumber + nobodys) % karts != 0):
            karts -= 1
        nobodys += 1
        if nobodys > 5:
            raise tooManyNobodysError
    if nobodys > 0:
        for i in range(nobodys):
            participantsNumber += 1
            participantsTab.append(str("Mr. Nobody %s") % str(i + 1))
    groups = participantsNumber // karts
    placesTab = list()
    for i in range(runsPerParticipant):
        if i % 2 == 0:
            currentRunTab = list(participantsTab)
            writtenParticipants = 0
            for f in range(groups):
                groupTab = list()
                for j in range(karts):
                    currentParticipant = currentRunTab[writtenParticipants]
                    groupTab.append(currentParticipant)
                    writtenParticipants += 1
                placesTab.append(groupTab)    
        elif i % 2 != 0:
            currentRunTab = [[karts * groups for groups in range(0, groups)] for karts in range(0, karts)]
            for f in range(groups):
                for j in range(karts):
                    currentRunTab[j][f] = str(placesTab[((i * groups) - groups) + f][j])
            for f in range(groups):
                currentRunTab[f].sort(reverse = revCalc)
            for f in range(groups):
                groupTab = list()
                for j in range(-1, (karts * -1) - 1, -1):
                    currentParticipant = currentRunTab[j][f]
                    groupTab.append(currentParticipant)
                placesTab.append(groupTab)
    parsePlaces(placesTab, pathToFile)